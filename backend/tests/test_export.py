"""Tests for Excel export endpoints (export.py)."""
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.financial import Account, AccountType, AccountDirection, VoucherStatus


class TestBalanceSheetExport:
    """Test balance sheet Excel export."""

    def test_export_balance_sheet_format(self, client, auth_headers, ledger, posted_voucher):
        """Balance sheet should return valid Excel with correct content-type."""
        resp = client.get(
            "/api/v1/export/balance-sheet?as_of_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "balance_sheet" in resp.headers["content-disposition"]

        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active
        assert ws.title == "资产负债表"
        # Title row should contain 资产负债表
        assert "资产负债表" in str(ws.cell(row=1, column=1).value)

    def test_balance_sheet_empty_period(self, client, auth_headers, ledger_headers):
        """Balance sheet for future date with no data returns empty but valid."""
        resp = client.get(
            "/api/v1/export/balance-sheet?as_of_date=2020-01-01",
            headers={**auth_headers, **ledger_headers},
        )
        assert resp.status_code == 200

    def test_balance_sheet_has_asset_totals(self, client, auth_headers, ledger, posted_voucher):
        """Posted voucher (bank debit 100k) should appear as asset."""
        resp = client.get(
            "/api/v1/export/balance-sheet?as_of_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active

        # Find "资产总计" row
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0] and "资产总计" in str(row[0]):
                return
        pytest.fail("Balance sheet should contain 资产总计")

    def test_balance_sheet_equation(self, client, auth_headers, ledger, posted_voucher):
        """Assets = Liabilities + Equity (accounting equation)."""
        resp = client.get(
            "/api/v1/export/balance-sheet?as_of_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active

        # Collect totals: columns A/B=assets, C/D=liabilities+equity
        total_assets = None
        total_le = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
            a_text = str(row[0].value) if row[0].value else ""
            c_text = str(row[2].value) if row[2].value else ""
            b_val = row[1].value
            d_val = row[3].value
            if "资产总计" in a_text:
                total_assets = float(b_val) if b_val else 0
            if "负债及权益总计" in c_text:
                total_le = float(d_val) if d_val else 0

        assert total_assets is not None
        assert total_le is not None
        assert abs(total_assets - total_le) < 0.01


class TestIncomeStatementExport:
    """Test income statement (利润表) Excel export."""

    def test_export_income_statement(self, client, auth_headers, ledger, revenue_voucher, expense_voucher):
        """Income statement for period with revenue and expense should show net income."""
        resp = client.get(
            "/api/v1/export/income-statement?start_date=2024-01-01&end_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active
        assert ws.title == "利润表"

        # Search for net income
        found_net_income = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0] and "净利润" in str(row[0]):
                found_net_income = True
                break
        assert found_net_income, "Income statement should contain 净利润"

    def test_income_statement_empty_period(self, client, auth_headers, ledger_headers):
        """Export for period with no transactions."""
        resp = client.get(
            "/api/v1/export/income-statement?start_date=2020-01-01&end_date=2020-12-31",
            headers={**auth_headers, **ledger_headers},
        )
        assert resp.status_code == 200

    def test_income_statement_has_revenue_section(self, client, auth_headers, ledger, revenue_voucher):
        """Revenue voucher (50k credit to 5001) should appear in revenue section."""
        resp = client.get(
            "/api/v1/export/income-statement?start_date=2024-01-01&end_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0] and "5001" in str(row[0]):
                return
        pytest.fail("Income statement should contain revenue account 5001")


class TestCashFlowExport:
    """Test cash flow statement (现金流量表) Excel export."""

    def test_export_cash_flow(self, client, auth_headers, ledger, posted_voucher):
        """Basic cash flow export should produce valid Excel."""
        resp = client.get(
            "/api/v1/export/cash-flow?start_date=2024-01-01&end_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active
        assert ws.title == "现金流量表"

    def test_cash_flow_has_sections(self, client, auth_headers, ledger, posted_voucher):
        """Cash flow should contain Operating, Investing, Financing sections."""
        resp = client.get(
            "/api/v1/export/cash-flow?start_date=2024-01-01&end_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active

        sections = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0]:
                s = str(row[0])
                if "经营活动" in s:
                    sections.add("operating")
                elif "投资活动" in s:
                    sections.add("investing")
                elif "筹资活动" in s:
                    sections.add("financing")
        assert len(sections) >= 2  # at least operating + one other

    def test_cash_flow_empty_period(self, client, auth_headers, ledger_headers):
        """Export for period with no cash transactions."""
        resp = client.get(
            "/api/v1/export/cash-flow?start_date=2020-01-01&end_date=2020-12-31",
            headers={**auth_headers, **ledger_headers},
        )
        assert resp.status_code == 200


class TestSubsidiaryLedgerExport:
    """Test subsidiary ledger (明细账) Excel export."""

    def test_export_subsidiary_ledger(self, client, auth_headers, ledger, posted_voucher):
        """Subsidiary ledger for 1002 (银行存款) should show the posted entry."""
        resp = client.get(
            "/api/v1/export/subsidiary-ledger?account_code=1002&start_date=2024-01-01&end_date=2024-12-31",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active
        assert "明细账" in ws.title
        # Title should mention the account
        assert "1002" in str(ws.cell(row=1, column=1).value)

        # Verify opening balance and total rows exist
        has_opening = False
        has_total = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
            if row[0] and "本期合计" in str(row[0]):
                has_total = True
            if row[1] and "期初余额" in str(row[1]):
                has_opening = True
        assert has_opening, "Subsidiary ledger should have 期初余额 row"
        assert has_total, "Subsidiary ledger should have 本期合计 row"

    def test_subsidiary_ledger_nonexistent_account(self, client, auth_headers, ledger_headers):
        """Non-existent account code should return 404."""
        resp = client.get(
            "/api/v1/export/subsidiary-ledger?account_code=9999&start_date=2024-01-01&end_date=2024-12-31",
            headers={**auth_headers, **ledger_headers},
        )
        assert resp.status_code == 404

    def test_subsidiary_ledger_empty_period(self, client, auth_headers, ledger, ledger_headers):
        """Subsidiary ledger for account with no transactions in period."""
        resp = client.get(
            "/api/v1/export/subsidiary-ledger?account_code=1001&start_date=2020-01-01&end_date=2020-12-31",
            headers={**auth_headers, **ledger_headers},
        )
        assert resp.status_code == 200


class TestGeneralLedgerExport:
    """Test general ledger (总账) Excel export."""

    def test_export_general_ledger(self, client, auth_headers, ledger, posted_voucher):
        """General ledger for a year should produce valid Excel with account rows."""
        resp = client.get(
            "/api/v1/export/general-ledger?year=2024",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active
        assert "总账" in ws.title

        # Verify headers exist
        headers = []
        for col in range(1, 7):
            v = ws.cell(row=4, column=col).value
            if v:
                headers.append(str(v))
        assert "科目代码" in headers
        assert "科目名称" in headers

    def test_general_ledger_empty_year(self, client, auth_headers, ledger_headers):
        """General ledger for a year with no data."""
        resp = client.get(
            "/api/v1/export/general-ledger?year=2020",
            headers={**auth_headers, **ledger_headers},
        )
        assert resp.status_code == 200

    def test_general_ledger_contains_account(self, client, auth_headers, ledger, posted_voucher):
        """Account 1002 (银行存款) should appear with transaction data."""
        resp = client.get(
            "/api/v1/export/general-ledger?year=2024",
            headers={**auth_headers, **{"X-Ledger-Id": str(ledger.id)}},
        )
        wb = load_workbook(BytesIO(resp.read()))
        ws = wb.active

        found = False
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0] and str(row[0]) == "1002":
                found = True
                break
        assert found, "General ledger should contain account 1002"


class TestExportAuth:
    """Test ledger header requirements for export endpoints."""

    def test_export_requires_ledger_header(self, client, auth_headers):
        """Export endpoints require X-Ledger-Id header."""
        resp = client.get(
            "/api/v1/export/balance-sheet?as_of_date=2024-12-31",
            headers=auth_headers,
        )
        assert resp.status_code == 400  # missing ledger_id

    def test_export_works_with_ledger_header(self, client, ledger_headers):
        """Export endpoints work with just ledger_id (no auth required)."""
        resp = client.get(
            "/api/v1/export/balance-sheet?as_of_date=2024-12-31",
            headers=ledger_headers,
        )
        assert resp.status_code == 200
