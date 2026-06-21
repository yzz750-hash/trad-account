from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, case
from datetime import date
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.routers.ledgers import get_ledger_id
from app.models.financial import (
    Account, Voucher, VoucherEntry, VoucherStatus,
    AccountDirection, AccountType,
)

router = APIRouter()

# Sanitize user-controlled values for HTTP header safety
import re as _re
def _safe_filename(s: str) -> str:
    return _re.sub(r'[\r\n\\/:*?"<>|]', '_', s)


# -- shared styles --
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=16)
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=10, color="666666")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
MONEY_FONT = Font(name="Consolas", size=10)
SECTION_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
TOTAL_FONT = Font(name="Microsoft YaHei", bold=True, size=11)

RMB_FORMAT = '#,##0.00_ ;[Red]-#,##0.00 '


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_body_cell(cell, is_money=False):
    cell.font = MONEY_FONT if is_money else BODY_FONT
    cell.border = THIN_BORDER
    if is_money:
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = RMB_FORMAT
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_title(ws, title, subtitle, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUBTITLE_FONT
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22


def _auto_width(ws, cols, min_w=10, max_w=42):
    for c in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, len(str(val)))
        width = min(max(max_len + 4, min_w), max_w)
        ws.column_dimensions[get_column_letter(c)].width = width


# ---------------- Balance Sheet Export ----------------

@router.get("/balance-sheet")
def export_balance_sheet(
    as_of_date: date,
    db: Session = Depends(get_db),
    ledger_id: int = Depends(get_ledger_id),
):
    accounts = db.query(Account).filter(
        Account.ledger_id == ledger_id, Account.is_active == True
    ).all()

    # Batch-fetch balance-sheet account balances
    bs_accounts = [a for a in accounts if a.account_type in (AccountType.ASSET, AccountType.LIABILITY, AccountType.EQUITY)]
    bs_ids = [a.id for a in bs_accounts]
    balances = {}
    if bs_ids:
        rows = (
            db.query(
                VoucherEntry.account_id,
                func.sum(case((VoucherEntry.direction == AccountDirection.DEBIT, VoucherEntry.amount), else_=0)).label("debit_sum"),
                func.sum(case((VoucherEntry.direction == AccountDirection.CREDIT, VoucherEntry.amount), else_=0)).label("credit_sum"),
            )
            .join(Voucher, VoucherEntry.voucher_id == Voucher.id)
            .filter(
                Voucher.ledger_id == ledger_id,
                VoucherEntry.account_id.in_(bs_ids),
                Voucher.voucher_date <= as_of_date,
                Voucher.status == VoucherStatus.POSTED,
            ).group_by(VoucherEntry.account_id).all()
        )
        balances = {r.account_id: (Decimal(str(r.debit_sum or 0)), Decimal(str(r.credit_sum or 0))) for r in rows}

    assets, liabilities, equity = [], [], []
    for account in bs_accounts:
        debit, credit = balances.get(account.id, (Decimal("0"), Decimal("0")))
        opening = Decimal(str(account.opening_balance))
        bal = (opening + debit - credit) if account.balance_direction == AccountDirection.DEBIT else (opening + credit - debit)
        if bal != 0:
            item = (f"{account.code} {account.name}", round(bal, 2))
            if account.account_type == AccountType.ASSET:
                assets.append(item)
            elif account.account_type == AccountType.LIABILITY:
                liabilities.append(item)
            else:
                equity.append(item)

    wb = Workbook()
    ws = wb.active
    ws.title = "资产负债表"

    total_assets = sum(a[1] for a in assets)
    total_liabilities = sum(l[1] for l in liabilities)
    total_equity = sum(e[1] for e in equity)

    _write_title(ws, "资产负债表 (Balance Sheet)", f"截止日期: {as_of_date.isoformat()}  |  币种: CNY", 4)

    headers = ["项目", "金额", "项目", "金额"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, 4)

    row = 5
    # Assets
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    sc = ws.cell(row=row, column=1, value="一、资产 (Assets)")
    sc.font = Font(name="Microsoft YaHei", bold=True, size=11)
    sc.fill = SECTION_FILL
    for cc in range(1, 3):
        ws.cell(row=row, column=cc).border = THIN_BORDER
    sc2 = ws.cell(row=row, column=3, value="一、负债 (Liabilities)")
    sc2.font = Font(name="Microsoft YaHei", bold=True, size=11)
    sc2.fill = SECTION_FILL
    for cc in range(3, 5):
        ws.cell(row=row, column=cc).border = THIN_BORDER
    row += 1

    max_rows = max(len(assets), len(liabilities) + len(equity) + 1)  # +1 for equity header
    for i in range(max_rows):
        # left: assets
        if i < len(assets):
            _style_body_cell(ws.cell(row=row, column=1, value=assets[i][0]))
            _style_body_cell(ws.cell(row=row, column=2, value=assets[i][1]), is_money=True)
        # right: liabilities then equity
        if i < len(liabilities):
            _style_body_cell(ws.cell(row=row, column=3, value=liabilities[i][0]))
            _style_body_cell(ws.cell(row=row, column=4, value=liabilities[i][1]), is_money=True)
        elif i == len(liabilities):
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            sc3 = ws.cell(row=row, column=3, value="二、所有者权益 (Equity)")
            sc3.font = Font(name="Microsoft YaHei", bold=True, size=11)
            sc3.fill = SECTION_FILL
            for cc in range(3, 5):
                ws.cell(row=row, column=cc).border = THIN_BORDER
        elif i - len(liabilities) - 1 < len(equity):
            eq_idx = i - len(liabilities) - 1
            _style_body_cell(ws.cell(row=row, column=3, value=equity[eq_idx][0]))
            _style_body_cell(ws.cell(row=row, column=4, value=equity[eq_idx][1]), is_money=True)
        row += 1

    # totals
    for cc in range(1, 5):
        ws.cell(row=row, column=cc).border = THIN_BORDER
        ws.cell(row=row, column=cc).font = TOTAL_FONT
    ws.cell(row=row, column=1, value="资产总计")
    ws.cell(row=row, column=2, value=round(total_assets, 2))
    ws.cell(row=row, column=2).number_format = RMB_FORMAT
    ws.cell(row=row, column=3, value="负债及权益总计")
    ws.cell(row=row, column=4, value=round(total_liabilities + total_equity, 2))
    ws.cell(row=row, column=4).number_format = RMB_FORMAT

    _auto_width(ws, 4, min_w=28)
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 20

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=balance_sheet_{as_of_date.isoformat()}.xlsx"})


# ---------------- Income Statement Export ----------------

@router.get("/income-statement")
def export_income_statement(
    start_date: date, end_date: date,
    db: Session = Depends(get_db),
    ledger_id: int = Depends(get_ledger_id),
):
    accounts = db.query(Account).filter(
        Account.ledger_id == ledger_id, Account.is_active == True,
        Account.account_type == AccountType.PROFIT_LOSS,
    ).all()

    # Batch-fetch all PL account balances for the period
    pl_ids = [a.id for a in accounts]
    balances = {}
    if pl_ids:
        rows = (
            db.query(
                VoucherEntry.account_id,
                func.sum(case((VoucherEntry.direction == AccountDirection.DEBIT, VoucherEntry.amount), else_=0)).label("debit_sum"),
                func.sum(case((VoucherEntry.direction == AccountDirection.CREDIT, VoucherEntry.amount), else_=0)).label("credit_sum"),
            )
            .join(Voucher, VoucherEntry.voucher_id == Voucher.id)
            .filter(
                Voucher.ledger_id == ledger_id,
                VoucherEntry.account_id.in_(pl_ids),
                Voucher.voucher_date >= start_date, Voucher.voucher_date <= end_date,
                Voucher.status == VoucherStatus.POSTED,
            ).group_by(VoucherEntry.account_id).all()
        )
        balances = {r.account_id: (Decimal(str(r.debit_sum or 0)), Decimal(str(r.credit_sum or 0))) for r in rows}

    revenues, expenses = [], []
    for account in accounts:
        debit, credit = balances.get(account.id, (Decimal("0"), Decimal("0")))
        if account.balance_direction == AccountDirection.DEBIT:
            net = debit - credit
            if net != 0:
                expenses.append((f"{account.code} {account.name}", round(net, 2)))
        else:
            net = credit - debit
            if net != 0:
                revenues.append((f"{account.code} {account.name}", round(net, 2)))

    total_revenue = sum(r[1] for r in revenues)
    total_expense = sum(e[1] for e in expenses)
    net_income = total_revenue - total_expense

    wb = Workbook()
    ws = wb.active
    ws.title = "利润表"

    _write_title(ws, "利润表 (Income Statement)", f"期间: {start_date.isoformat()} ~ {end_date.isoformat()}  |  币种: CNY", 2)

    headers = ["项目", "金额"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, 2)

    row = 5
    def _section(title, items, subtotal):
        nonlocal row
        sc = ws.cell(row=row, column=1, value=title)
        sc.font = Font(name="Microsoft YaHei", bold=True, size=11)
        sc.fill = SECTION_FILL
        ws.cell(row=row, column=2).fill = SECTION_FILL
        for cc in range(1, 3):
            ws.cell(row=row, column=cc).border = THIN_BORDER
        row += 1
        for name, amt in items:
            _style_body_cell(ws.cell(row=row, column=1, value=name))
            _style_body_cell(ws.cell(row=row, column=2, value=amt), is_money=True)
            row += 1
        for cc in range(1, 3):
            ws.cell(row=row, column=cc).border = THIN_BORDER
            ws.cell(row=row, column=cc).font = TOTAL_FONT
        ws.cell(row=row, column=1, value="小计")
        ws.cell(row=row, column=2, value=round(subtotal, 2))
        ws.cell(row=row, column=2).number_format = RMB_FORMAT
        row += 1

    _section("一、营业收入 (Revenue)", revenues, total_revenue)
    row += 1
    _section("二、营业支出 (Expenses)", expenses, total_expense)
    row += 1

    for cc in range(1, 3):
        ws.cell(row=row, column=cc).border = THIN_BORDER
        ws.cell(row=row, column=cc).font = Font(name="Microsoft YaHei", bold=True, size=13, color="1F2937")
    ws.cell(row=row, column=1, value="净利润 (Net Income)")
    ws.cell(row=row, column=2, value=round(net_income, 2))
    ws.cell(row=row, column=2).number_format = RMB_FORMAT
    ws.row_dimensions[row].height = 28

    _auto_width(ws, 2, min_w=36)
    ws.column_dimensions["B"].width = 22

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=income_statement_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"})


# ---------------- Cash Flow Statement Export ----------------

@router.get("/cash-flow")
def export_cash_flow(
    start_date: date, end_date: date,
    db: Session = Depends(get_db),
    ledger_id: int = Depends(get_ledger_id),
):
    from sqlalchemy.orm import joinedload

    cash_accounts = db.query(Account).filter(
        Account.ledger_id == ledger_id,
        Account.code.like('1001%') | Account.code.like('1002%')
    ).all()
    cash_account_ids = {a.id for a in cash_accounts}

    operating_inflows, operating_outflows = {}, {}
    investing_inflows, investing_outflows = {}, {}
    financing_inflows, financing_outflows = {}, {}

    ending_cash_balance = Decimal("0")

    if cash_account_ids:
        cash_vouchers = (
            db.query(Voucher)
            .options(joinedload(Voucher.entries).joinedload(VoucherEntry.account))
            .join(VoucherEntry, Voucher.id == VoucherEntry.voucher_id)
            .filter(
                Voucher.ledger_id == ledger_id,
                VoucherEntry.account_id.in_(cash_account_ids),
                Voucher.voucher_date >= start_date, Voucher.voucher_date <= end_date,
                Voucher.status == VoucherStatus.POSTED,
            ).distinct().all()
        )

        for v in cash_vouchers:
            cash_entries = [e for e in v.entries if e.account_id in cash_account_ids]
            non_cash_entries = [e for e in v.entries if e.account_id not in cash_account_ids]
            if not non_cash_entries:
                continue
            total_cash_inflow = sum(e.amount for e in cash_entries if e.direction == AccountDirection.DEBIT)
            total_cash_outflow = sum(e.amount for e in cash_entries if e.direction == AccountDirection.CREDIT)
            net_impact = Decimal(str(total_cash_inflow - total_cash_outflow))
            if abs(net_impact) < Decimal("0.01"):
                continue
            primary = max(non_cash_entries, key=lambda e: e.amount)
            code = primary.account.code
            name = primary.account.name
            is_inflow = net_impact > 0
            amt = abs(net_impact)
            if code.startswith('16'):
                target = investing_inflows if is_inflow else investing_outflows
            elif code.startswith('20') or code.startswith('4'):
                target = financing_inflows if is_inflow else financing_outflows
            else:
                target = operating_inflows if is_inflow else operating_outflows
            target[name] = target.get(name, 0) + amt

        # Batch-fetch ending balances for all cash accounts
        if cash_account_ids:
            cash_balance_rows = (
                db.query(
                    VoucherEntry.account_id,
                    func.sum(case((VoucherEntry.direction == AccountDirection.DEBIT, VoucherEntry.amount), else_=0)).label("debit_sum"),
                    func.sum(case((VoucherEntry.direction == AccountDirection.CREDIT, VoucherEntry.amount), else_=0)).label("credit_sum"),
                )
                .join(Voucher, VoucherEntry.voucher_id == Voucher.id)
                .filter(Voucher.ledger_id == ledger_id, VoucherEntry.account_id.in_(cash_account_ids),
                        Voucher.voucher_date <= end_date,
                        Voucher.status == VoucherStatus.POSTED).group_by(VoucherEntry.account_id).all()
            )
            cash_balance_map = {r.account_id: (Decimal(str(r.debit_sum or 0)), Decimal(str(r.credit_sum or 0))) for r in cash_balance_rows}
            for acc in cash_accounts:
                debit, credit = cash_balance_map.get(acc.id, (Decimal("0"), Decimal("0")))
                ending_cash_balance += Decimal(str(acc.opening_balance)) + debit - credit

    def _write_section(ws, start_row, title, inflows, outflows):
        r = start_row
        sc = ws.cell(row=r, column=1, value=title)
        sc.font = Font(name="Microsoft YaHei", bold=True, size=11)
        sc.fill = SECTION_FILL
        for cc in range(1, 3):
            ws.cell(row=r, column=cc).fill = SECTION_FILL
            ws.cell(row=r, column=cc).border = THIN_BORDER
        r += 1
        subtotal = 0
        if inflows:
            ws.cell(row=r, column=1, value="  现金流入")
            ws.cell(row=r, column=1).font = Font(name="Microsoft YaHei", bold=True, size=10, color="059669")
            r += 1
            for n, a in inflows.items():
                _style_body_cell(ws.cell(row=r, column=1, value=f"    {n}"))
                _style_body_cell(ws.cell(row=r, column=2, value=round(a, 2)), is_money=True)
                subtotal += a
                r += 1
        if outflows:
            ws.cell(row=r, column=1, value="  现金流出")
            ws.cell(row=r, column=1).font = Font(name="Microsoft YaHei", bold=True, size=10, color="DC2626")
            r += 1
            for n, a in outflows.items():
                _style_body_cell(ws.cell(row=r, column=1, value=f"    {n}"))
                _style_body_cell(ws.cell(row=r, column=2, value=round(a, 2)), is_money=True)
                subtotal -= a
                r += 1
        # subtotal row
        for cc in range(1, 3):
            ws.cell(row=r, column=cc).border = THIN_BORDER
            ws.cell(row=r, column=cc).font = TOTAL_FONT
        ws.cell(row=r, column=1, value="净额")
        ws.cell(row=r, column=2, value=round(subtotal, 2))
        ws.cell(row=r, column=2).number_format = RMB_FORMAT
        r += 1
        return r, subtotal

    wb = Workbook()
    ws = wb.active
    ws.title = "现金流量表"
    _write_title(ws, "现金流量表 (Cash Flow Statement)",
                 f"期间: {start_date.isoformat()} ~ {end_date.isoformat()}  |  币种: CNY", 2)
    headers = ["项目", "金额"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, 2)

    row = 5
    row, net_op = _write_section(ws, row, "一、经营活动 (Operating)", operating_inflows, operating_outflows)
    row += 1
    row, net_inv = _write_section(ws, row, "二、投资活动 (Investing)", investing_inflows, investing_outflows)
    row += 1
    row, net_fin = _write_section(ws, row, "三、筹资活动 (Financing)", financing_inflows, financing_outflows)
    row += 1

    net_increase = net_op + net_inv + net_fin
    for metric_name, metric_val in [("本期现金净增加额", net_increase), ("期末现金余额", ending_cash_balance)]:
        for cc in range(1, 3):
            ws.cell(row=row, column=cc).border = THIN_BORDER
            ws.cell(row=row, column=cc).font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F2937")
        ws.cell(row=row, column=1, value=metric_name)
        ws.cell(row=row, column=2, value=round(metric_val, 2))
        ws.cell(row=row, column=2).number_format = RMB_FORMAT
        ws.row_dimensions[row].height = 26
        row += 1

    _auto_width(ws, 2, min_w=36)
    ws.column_dimensions["B"].width = 22

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=cash_flow_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"})


# ---------------- Subsidiary Ledger Export ----------------

@router.get("/subsidiary-ledger")
def export_subsidiary_ledger(
    account_code: str, start_date: date, end_date: date,
    db: Session = Depends(get_db),
    ledger_id: int = Depends(get_ledger_id),
):
    account = db.query(Account).filter(
        Account.ledger_id == ledger_id, Account.code == account_code
    ).first()
    if not account:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Account not found")

    entries_query = (
        db.query(VoucherEntry, Voucher)
        .join(Voucher, VoucherEntry.voucher_id == Voucher.id)
        .filter(
            Voucher.ledger_id == ledger_id,
            VoucherEntry.account_id == account.id,
            Voucher.voucher_date >= start_date, Voucher.voucher_date <= end_date,
            Voucher.status == VoucherStatus.POSTED,
        ).order_by(Voucher.voucher_date.asc(), Voucher.id.asc()).all()
    )

    past = (
        db.query(
            func.sum(case((VoucherEntry.direction == AccountDirection.DEBIT, VoucherEntry.amount), else_=0)).label("debit_sum"),
            func.sum(case((VoucherEntry.direction == AccountDirection.CREDIT, VoucherEntry.amount), else_=0)).label("credit_sum"),
        )
        .join(Voucher, VoucherEntry.voucher_id == Voucher.id)
        .filter(Voucher.ledger_id == ledger_id, VoucherEntry.account_id == account.id,
                Voucher.voucher_date < start_date,
                Voucher.status == VoucherStatus.POSTED).first()
    )
    past_debit = Decimal(str(past.debit_sum or 0))
    past_credit = Decimal(str(past.credit_sum or 0))
    opening = Decimal(str(account.opening_balance))

    if account.balance_direction == AccountDirection.DEBIT:
        balance = opening + past_debit - past_credit
    else:
        balance = opening + past_credit - past_debit

    wb = Workbook()
    ws = wb.active
    ws.title = "明细账"

    _write_title(ws, f"明细账 (Subsidiary Ledger) — {account.code} {account.name}",
                 f"期间: {start_date.isoformat()} ~ {end_date.isoformat()}  |  币种: CNY", 6)

    headers = ["日期", "凭证字号", "摘要", "借方金额", "贷方金额", "余额"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, 6)

    row = 5
    # opening balance
    for cc in range(1, 7):
        _style_body_cell(ws.cell(row=row, column=cc))
    ws.cell(row=row, column=1, value=str(start_date))
    ws.cell(row=row, column=2, value="期初余额")
    ws.cell(row=row, column=3, value="期初余额")
    ws.cell(row=row, column=6, value=round(balance, 2))
    ws.cell(row=row, column=6).number_format = RMB_FORMAT
    row += 1

    total_debit = Decimal("0")
    total_credit = Decimal("0")
    for entry, voucher in entries_query:
        debit = Decimal(str(entry.amount)) if entry.direction == AccountDirection.DEBIT else Decimal("0")
        credit = Decimal(str(entry.amount)) if entry.direction == AccountDirection.CREDIT else Decimal("0")
        if account.balance_direction == AccountDirection.DEBIT:
            balance = balance + debit - credit
        else:
            balance = balance + credit - debit
        total_debit += debit
        total_credit += credit

        _style_body_cell(ws.cell(row=row, column=1, value=str(voucher.voucher_date)))
        _style_body_cell(ws.cell(row=row, column=2, value=voucher.voucher_number))
        _style_body_cell(ws.cell(row=row, column=3, value=entry.summary))
        _style_body_cell(ws.cell(row=row, column=4, value=debit if debit else None), is_money=True)
        _style_body_cell(ws.cell(row=row, column=5, value=credit if credit else None), is_money=True)
        _style_body_cell(ws.cell(row=row, column=6, value=round(balance, 2)), is_money=True)
        row += 1

    # totals
    for cc in range(1, 7):
        ws.cell(row=row, column=cc).border = THIN_BORDER
        ws.cell(row=row, column=cc).font = TOTAL_FONT
    ws.cell(row=row, column=1, value="本期合计")
    ws.cell(row=row, column=4, value=round(total_debit, 2))
    ws.cell(row=row, column=4).number_format = RMB_FORMAT
    ws.cell(row=row, column=5, value=round(total_credit, 2))
    ws.cell(row=row, column=5).number_format = RMB_FORMAT
    ws.cell(row=row, column=6, value=round(balance, 2))
    ws.cell(row=row, column=6).number_format = RMB_FORMAT

    _auto_width(ws, 6, min_w=10)
    ws.column_dimensions["C"].width = 36

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=subsidiary_{_safe_filename(account_code)}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"})


# ---------------- General Ledger Export ----------------

@router.get("/general-ledger")
def export_general_ledger(
    year: int,
    db: Session = Depends(get_db),
    ledger_id: int = Depends(get_ledger_id),
):
    accounts = db.query(Account).filter(
        Account.ledger_id == ledger_id, Account.is_active == True
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "总账"

    _write_title(ws, f"总账 (General Ledger) — {year}年度", f"币种: CNY", 6)

    headers = ["科目代码", "科目名称", "月份", "借方合计", "贷方合计", "余额"]

    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, 6)

    # Batch-fetch past (pre-year) and monthly balances for all accounts
    account_ids = [a.id for a in accounts]
    past_map = {}
    monthly_all = {}
    if account_ids:
        past_rows = (
            db.query(
                VoucherEntry.account_id,
                func.sum(case((VoucherEntry.direction == AccountDirection.DEBIT, VoucherEntry.amount), else_=0)).label("debit_sum"),
                func.sum(case((VoucherEntry.direction == AccountDirection.CREDIT, VoucherEntry.amount), else_=0)).label("credit_sum"),
            )
            .join(Voucher, VoucherEntry.voucher_id == Voucher.id)
            .filter(
                Voucher.ledger_id == ledger_id,
                VoucherEntry.account_id.in_(account_ids),
                extract("year", Voucher.voucher_date) < year,
                Voucher.status == VoucherStatus.POSTED,
            ).group_by(VoucherEntry.account_id).all()
        )
        past_map = {r.account_id: (Decimal(str(r.debit_sum or 0)), Decimal(str(r.credit_sum or 0))) for r in past_rows}

        monthly_rows = (
            db.query(
                VoucherEntry.account_id,
                extract("month", Voucher.voucher_date).label("month"),
                func.sum(case((VoucherEntry.direction == AccountDirection.DEBIT, VoucherEntry.amount), else_=0)).label("debit_sum"),
                func.sum(case((VoucherEntry.direction == AccountDirection.CREDIT, VoucherEntry.amount), else_=0)).label("credit_sum"),
            )
            .join(Voucher, VoucherEntry.voucher_id == Voucher.id)
            .filter(
                Voucher.ledger_id == ledger_id,
                VoucherEntry.account_id.in_(account_ids),
                extract("year", Voucher.voucher_date) == year,
                Voucher.status == VoucherStatus.POSTED,
            ).group_by(VoucherEntry.account_id, extract("month", Voucher.voucher_date)).all()
        )
        for r in monthly_rows:
            monthly_all.setdefault(r.account_id, {})[int(r.month)] = (Decimal(str(r.debit_sum or 0)), Decimal(str(r.credit_sum or 0)))

    row = 5
    for account in accounts:
        past_debit, past_credit = past_map.get(account.id, (Decimal("0"), Decimal("0")))
        opening = Decimal(str(account.opening_balance))
        balance = (opening + past_debit - past_credit) if account.balance_direction == AccountDirection.DEBIT else (opening + past_credit - past_debit)

        month_dict = monthly_all.get(account.id, {})

        has_data = False
        for m in range(1, 13):
            d, c = month_dict.get(m, (Decimal("0"), Decimal("0")))
            if d == 0 and c == 0 and balance == 0:
                continue
            has_data = True
            if account.balance_direction == AccountDirection.DEBIT:
                balance = balance + d - c
            else:
                balance = balance + c - d
            _style_body_cell(ws.cell(row=row, column=1, value=account.code))
            _style_body_cell(ws.cell(row=row, column=2, value=account.name))
            _style_body_cell(ws.cell(row=row, column=3, value=f"{year}-{m:02d}"))
            _style_body_cell(ws.cell(row=row, column=4, value=d if d else None), is_money=True)
            _style_body_cell(ws.cell(row=row, column=5, value=c if c else None), is_money=True)
            _style_body_cell(ws.cell(row=row, column=6, value=round(balance, 2)), is_money=True)
            row += 1

        if not has_data and balance != 0:
            _style_body_cell(ws.cell(row=row, column=1, value=account.code))
            _style_body_cell(ws.cell(row=row, column=2, value=account.name))
            _style_body_cell(ws.cell(row=row, column=3, value="期初"))
            _style_body_cell(ws.cell(row=row, column=6, value=round(balance, 2)), is_money=True)
            row += 1

    _auto_width(ws, 6, min_w=12)
    ws.column_dimensions["B"].width = 28

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=general_ledger_{year}.xlsx"})
